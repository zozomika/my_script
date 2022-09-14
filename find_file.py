#!/common/appl/python/python-2.7.13-64bit/bin/python

import sys, time, datetime
from openpyxl import load_workbook, Workbook
from openpyxl.styles import PatternFill, Border, Side, Alignment, Protection, Font, colors, NamedStyle
from openpyxl.comments import Comment
from openpyxl.comments.comment_sheet import Properties
from subprocess import Popen, PIPE
import re, os, glob, fnmatch
 
t0 = time.time()
 
wb1 = load_workbook(sys.argv[1])
 
# Setting
ignore_sheet = ["Rule","E3_History", "Delete(E3)","Old_V3M_History","hard_macro (MIPI)","D3_History"]
ignore_cell = ["Use","use"]
 
# Initial declaration
wb1_sheets = wb1.sheetnames

# get 1st match
def find(name, path):
    for root, dirs, files in os.walk(path):
        if name in files:
            return os.path.join(root, name)
#get all matches
def find_all(name, path):
    result = []
    for root, dirs, files in os.walk(path):
        if name in files:
            result.append(os.path.join(root, name))
    return result
#get match by patterns:
def find_pattern (pattern, path):
    result = []
    for root, dirs, files in os.walk(path):
        for name in files:
            #if str.lower(name) in [x.lower() for x in files]:
            if fnmatch.fnmatch(name, pattern) :
                result.append(os.path.join(root, name))
    return result

#find_pattern('*.gds', r'/design01/rcarg3_lib2/TSMC_28HPC/core_cell/9T/100A/Renesas_Customized/20161129/L_T28HPCBWP40P140LVT_V01.00.01/gds/T28HPCBWP40P140LVT/')
#a = "/design01/rcarg3_lib2/TSMC_28HPC/core_cell/9T/100A/Renesas_Customized/20161129/L_T28HPCBWP40P140LVT_V01.00.01/gds/T28HPCBWP40P140LVT/"
#import glob
#for x in glob.glob(a+'*.gds'):
#    print(x)

for s1 in wb1_sheets: 
        if s1 == 'GDSMERGE_1':
                max_row = wb1[s1].max_row
                max_col = wb1[s1].max_column
                for r_cnt in range(1,max_row+1):
                        for c_cnt in range(1, max_col+1):
                                #if wb0[s1].cell(row=r_cnt, column=c_cnt).value == 'gds':
                                        
                                line_new =  wb1[s1].cell(row=r_cnt, column=c_cnt).value
                                if line_new:
                                        if re.match (r'.*\.gds\s*$',line_new):
                                                #print line_new
                                                line_new_mdfy = re.sub(r'\s*GLIB NAME =\s*','',line_new)
                                                line_new_mdfy = line_new_mdfy.strip()
                                                print line_new_mdfy
                                        
                                        
        
