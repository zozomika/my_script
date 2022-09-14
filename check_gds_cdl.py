#!/common/appl/python/python-2.7.13-64bit/bin/python
#.---------------- Author: HauHuynh
#.---------------- Function: Get gds, cdl, nwpg from library management sheet (LMS) and Verification Result (VR)
#.---------------- Usage: script.py <lib_excel_file> <verification_result>
#.---------------- Ver00: + Can get cdl, gds, nwpg raw path from LMS - Oct 21, 2017
#                         + Restriction: can not get included cell's libraries (Ex:[CDL] *source.added)
#.---------------- Ver01: + Can ignore unused sheets, check used only input - Nov 3, 2017
#                         + Can get all cdl, gds, nwpg files 
#.---------------- Ver02: Can get get cdl, gds, nwpg from VR
#.---------------- Ver03: Can get get cdl included from LMS cdl files - Nov 7, 2017
#.---------------- Ver04: to be updated - should be able to ignore hidden sheet
import sys, time, datetime
from openpyxl import load_workbook, Workbook
from openpyxl.styles import PatternFill, Border, Side, Alignment, Protection, Font, colors, NamedStyle
from openpyxl.comments import Comment
from openpyxl.comments.comment_sheet import Properties
import re, os,  fnmatch

t0 = time.time()
t2 = time.time()
wb0 = load_workbook(sys.argv[1]) # LMS
wb1 = load_workbook(sys.argv[2]) # VR
t3 = time.time()
print "Time to load workbook: %s" %(t3-t2)
output = Workbook()

# Setting
#ignore_sheet = ["Rule","E3_History", "Delete(E3)","Old_V3M_History","hard_macro (MIPI)","D3_History", "Delete(D3)"]
ignore_sheet = ["Rule","E3_History", "Delete(E3)","Old_V3M_History","hard_macro (MIPI)","D3_History", "Delete(D3)"]
ignore_cell = ["Use","use"]

# Initial declaration

stamp_out = '{:%y%m%d_%H%M}'.format(datetime.datetime.now()) #Ex: 171103_1234
# For LMS
wb0_sheets = wb0.sheetnames
gds_list = []
cdl_list = []
nwpg_list = []
# --Output File
gds_out = "fr_LMS_gds_" + stamp_out + ".txt"
cdl_out = "fr_LMS_cdl_" + stamp_out + ".txt"
nwpg_out = "fr_LMS_nwpg_" + stamp_out + ".txt"
gds = open(gds_out, 'w')
cdl = open(cdl_out, 'w')
nwpg = open(nwpg_out, 'w')

# For VR
wb1_sheets = wb1.sheetnames
gdsmerge_list = []
cdlcat_list = []
nwpg_list_vr = []

# --Output File
gds_verify = "fr_Verification_report_gds_" + stamp_out + ".txt"   
cdl_verify = "fr_Verification_report_cdl_" + stamp_out + ".txt"   
nw_verify = "fr_Verification_report_nwpg_" + stamp_out + ".txt"

gdsmerge = open(gds_verify, 'w')
cdlcat = open(cdl_verify, 'w')
nwpg_vr = open(nw_verify,'w')

# Define function to get file by an address
def find_pattern (pattern, path):
    result = []
    for root, dirs, files in os.walk(path):
        for name in files:
            if fnmatch.fnmatch(name, pattern) :
                result.append(os.path.join(root, name))
    return result
# Below function has riskly to use
#def find_pattern (pattern, path):
#        result = []
#        for file in glob.glob(path+pattern):
#                result.append(file)
#        print result

# Define function to write output file by list
def write_out (list_write_out, output):
  for x in sorted(set(list_write_out)):
    output.write(x)

#find_pattern ('*.gds$','/design02/rcarg3_lib5/TSMC_28HPC/core_cell/7T/TSMC_Original/V100_20170426/Back_End/gds/tcbn28hpcbwp7t35p140lvt_100c/')

# Analysing LMS
print "[LMS] Reading data from file %s   z.Z.Z..." % os.path.basename(sys.argv[1])
for s0 in wb0_sheets:
  if s0 in ignore_sheet:
     print "[Warning!][LMS] Ignore excel sheet:", s0
     continue
  else:
     print "[LMS] Reading excel sheet:", s0
     #wb0_sheets.remove(s0) 
     max_row = wb0[s0].max_row
     max_col = wb0[s0].max_column
     for r_cnt in range(1,max_row+1):
       for c_cnt in range(1, max_col+1):
          if wb0[s0].cell(row=r_cnt, column=c_cnt).value == 'gds': #GDS Part
             if wb0[s0].cell(row=r_cnt, column=c_cnt-1).value in ignore_cell and wb0[s0].cell(row=r_cnt, column=c_cnt+1).value:
                gds_path = wb0[s0].cell(row=r_cnt, column=c_cnt+1).value
                if os.path.isdir(gds_path): # check if gds_path is a directory addess
                #gds.write(gds_path)
                   for file in find_pattern('*.gds*', gds_path):
                      gds_list.append(file+"\n")
                      #gds.write(file+"\n")
                elif fnmatch.fnmatch(gds_path, '*.gds*'): # check if gds_path is a file address
                      #gds.write(gds_path+"\n")
                      gds_list.append(gds_path+"\n")
          elif wb0[s0].cell(row=r_cnt, column=c_cnt).value == 'cdl': #CDL and NWPG part
             if wb0[s0].cell(row=r_cnt, column=c_cnt-1).value in ignore_cell and wb0[s0].cell(row=r_cnt, column=c_cnt+1).value:
                cdl_path = wb0[s0].cell(row=r_cnt, column=c_cnt+1).value
                if os.path.isdir(cdl_path):
                   for file in find_pattern('*.cdl', cdl_path):
                      #cdl.write(file+"\n")
                      cdl_list.append(file+"\n")
                      origin_file = open(file,'r')
                      for line in origin_file.readlines():
                         #line = re.findall(r'.INCLUDE', line)
                         if r'.INCLUDE' in line and r'*.INCLUDE' not in line:
                            line = line.split('INCLUDE')[1]
                            line = line.strip()
                            #if line not in cdl_list:
                            cdl_list.append(line+"\n")
                      origin_file.close()
                   for file in find_pattern('*.spi', cdl_path):
                      #cdl.write(file+"\n")
                      cdl_list.append(file+"\n")
                   #for file in find_pattern('*source.added', cdl_path):
                   #   cdl_list.append(file+"\n")
                elif fnmatch.fnmatch(cdl_path, '*.spi') or fnmatch.fnmatch(cdl_path, '*.cdl'):
                   #cdl.write(cdl_path+"\n")
                   cdl_list.append(cdl_path+"\n")
          #         cdl.write(wb0[s0].cell(row=r_cnt, column=c_cnt+1).value + "\n")
          elif wb0[s0].cell(row=r_cnt, column=c_cnt).value == 'netwalker(pg)':
             if wb0[s0].cell(row=r_cnt, column=c_cnt-1).value in ignore_cell and wb0[s0].cell(row=r_cnt, column=c_cnt+1).value:           
                nwpg_path = wb0[s0].cell(row=r_cnt, column=c_cnt+1).value
                if os.path.isdir(nwpg_path):
                   for file in find_pattern('*.nw', nwpg_path):
                      #nwpg.write(file+"\n")
                      nwpg_list.append(file+"\n")
                elif fnmatch.fnmatch(nwpg_path, '*.nw'):
                   #nwpg.write(nwpg_path+"\n")
                   nwpg_list.append(nwpg_path+"\n")
          #        nwpg.write(wb0[s0].cell(row=r_cnt, column=c_cnt+1).value + "\n")
                   #wb1[sheet].cell(row=r_cnt, column=c_cnt).style = highlight

#Analysing Verification sheet:
print "[VR] Reading data from file %s   z.Z.Z..." % os.path.basename(sys.argv[2])
for s1 in wb1_sheets:
   if s1 == 'GDSMERGE_1':
      print "[VR] Reading excel sheet:", s1
      max_row = wb1[s1].max_row
      max_col = wb1[s1].max_column
      for r_cnt in range(1,max_row+1):
         for c_cnt in range(1, max_col+1):
            #if wb0[s1].cell(row=r_cnt, column=c_cnt).value == 'gds':
            line_new =  wb1[s1].cell(row=r_cnt, column=c_cnt).value
            if isinstance(line_new, basestring) and line_new:
               line_new = line_new.strip()
               #line_new1 = [elem.strip() for elem in line_new]
               #if re.match (r'.*GLIB NAME =.*\.gds.*$',line_new) and re.match(r'.*GLIB NAME =.*\.gdsii.*$',line_new) is False:
               if re.match (r'.*GLIB NAME =.*\.gds.*$',line_new) :
                  line_new_mdfy = re.sub(r'\s*GLIB NAME =\s*','',line_new)
                  #line_new_mdfy = re.sub(r'.*\/\/.*','\/',line_new_mdfy)
                  line_new_mdfy = line_new_mdfy.replace("//","/") 
                  gdsmerge_list.append(line_new_mdfy+"\n")
                  #print line_new_mdfy

   elif s1 == 'CDLCAT':
      print "[VR] Reading excel sheet:", s1
      max_row = wb1[s1].max_row
      max_col = wb1[s1].max_column
      for r_cnt in range(1,max_row+1):
         for c_cnt in range(1, max_col+1):
            #if wb0[s1].cell(row=r_cnt, column=c_cnt).value == 'gds':
            line_new =  wb1[s1].cell(row=r_cnt, column=c_cnt).value
            if isinstance(line_new, basestring) and line_new:
               line_new = line_new.strip()
               if re.match (r'.*CDL :.*\.cdl.*$',line_new) or re.match (r'.*CDL :.*\.spi.*$',line_new) or re.match (r'.*CDL :.*source.added.*$',line_new):
                  line_new_mdfy = re.sub(r'\s*CDL :\s*','',line_new)
                  line_new_mdfy = line_new_mdfy.replace("//","/") 
                  #line_new_mdfy = re.sub(r'.*\/\/.*','\/',line_new_mdfy)
                  cdlcat_list.append(line_new_mdfy+"\n")
               if re.match ('.*LVSNET\/run.*',line_new):
                  nwpg_rp = re.sub(r'\/run.*$','/NW/lib.nwcell__hier',line_new)
                  if os.path.isfile(nwpg_rp):
                     print "[VR] Reading file:", nwpg_rp
                     nw_lib_src = open(nwpg_rp,'r')
                     nw_list = nw_lib_src.readlines()
                     for line in nw_list:
                        line_new = line.strip()
                        line_new_mdfy = line_new.replace("//","/")
                        line_new_mdfy = re.sub(r'\s*-NWCELL\s*','',line_new)
                        #print line_new_mdfy
                        nwpg_list_vr.append(line_new_mdfy+"\n")
                  else:
                     print "[ERROR][VR] The file %s you are trying to access is no longer available" % nwpg_rp
# Write out put
# --For VR
write_out (gdsmerge_list, gdsmerge)
write_out (cdlcat_list, cdlcat)
if os.path.isfile(nwpg_rp):
   write_out (nwpg_list_vr, nwpg_vr)
## --For LMS
write_out (gds_list, gds)
write_out (cdl_list, cdl)
write_out (nwpg_list, nwpg)


# Close output
# --For Vr
gdsmerge.close()                                                                                                                                                            
cdlcat.close()
nw_lib_src.close()
nwpg_vr.close()

# --For LMS
gds.close()
cdl.close()
nwpg.close()


t1 =  time.time()
print "Execution time: ", t1-t3

#eof
