#!/common/appl/python/python-2.7.13-64bit/bin/python
#.---------------- Author: HauHuynh
#.---------------- Function: Get gds, cdl, nwpg from library management sheet (LMS)
#.---------------- Usage: script.py <lib_excel_file>
#.---------------- Ver00: Can get cdl, gds, nwpg raw path by LM - Oct 21, 2017
#.---------------- Ver01: + Can ignore unused sheets, check used only input - Nov 3, 2017
#                         + Can get all cdl, gds, nwpg files 
#.---------------- Ver02: to be updated - should be able to ignore hidden sheet
import sys, time, datetime
from openpyxl import load_workbook, Workbook
from openpyxl.styles import PatternFill, Border, Side, Alignment, Protection, Font, colors, NamedStyle
from openpyxl.comments import Comment
from openpyxl.comments.comment_sheet import Properties
from subprocess import Popen, PIPE
import re, os, glob, fnmatch

t0 = time.time()

wb0 = load_workbook(sys.argv[1])
wb1 = load_workbook(sys.argv[2])
output = Workbook()

# Setting
ignore_sheet = ["Rule","E3_History", "Delete(E3)","Old_V3M_History","hard_macro (MIPI)","D3_History"]
ignore_cell = ["Use","use"]

# Initial declaration
wb0_sheets = wb0.sheetnames
wb1_sheets = wb1.sheetnames

#gds = open('fr_LMS_gds_all', 'w')
#cdl = open('fr_LMS_cdl_all', 'w')
#nwpg = open('fr_LMS_nwpg_all', 'w')
stamp_out = '{:%y%m%d_%H%M}'.format(datetime.datetime.now())
print stamp_out
gds_out = "fr_LMS_gds_" + stamp_out + ".txt"
cdl_out = "fr_LMS_cdl_" + stamp_out + ".txt"
nwpg_out = "fr_LMS_nwpg_" + stamp_out + ".txt"
gds = open(gds_out, 'w')
cdl = open(cdl_out, 'w')
nwpg = open(nwpg_out, 'w')
gds_list = []
cdl_list = []
nwpg_list = []

# Define function to get file by an address
def find_pattern (pattern, path):
    result = []
    for root, dirs, files in os.walk(path):
        for name in files:
            if fnmatch.fnmatch(name, pattern) :
                result.append(os.path.join(root, name))
    return result

#def find_pattern (pattern, path):
#        result = []
#        for file in glob.glob(path+pattern):
#                result.append(file)
#        print result

#find_pattern ('*.gds$','/design02/rcarg3_lib5/TSMC_28HPC/core_cell/7T/TSMC_Original/V100_20170426/Back_End/gds/tcbn28hpcbwp7t35p140lvt_100c/')

# Analysing LMS
for s0 in wb0_sheets:
        if s0 in ignore_sheet:
                continue
        else:
                #wb0_sheets.remove(s0) 
                max_row = wb0[s0].max_row
                max_col = wb0[s0].max_column
                for r_cnt in range(1,max_row+1):
                        for c_cnt in range(1, max_col+1):
                                if wb0[s0].cell(row=r_cnt, column=c_cnt).value == 'gds':
                                        if wb0[s0].cell(row=r_cnt, column=c_cnt-1).value in ignore_cell and wb0[s0].cell(row=r_cnt, column=c_cnt+1).value:
                                                gds_path = wb0[s0].cell(row=r_cnt, column=c_cnt+1).value
                                                if os.path.isdir(gds_path):
                                                #gds.write(gds_path)
                                                        for file in find_pattern('*.gds*', gds_path):
                                                                gds_list.append(file+"\n")
                                                                #gds.write(file+"\n")
                                                elif glob.glob(gds_path+'*.gds*') or fnmatch.fnmatch(gds_path, '*.gds*'):
                                                        #gds.write(gds_path+"\n")
                                                        gds_list.append(gds_path+"\n")
                                elif wb0[s0].cell(row=r_cnt, column=c_cnt).value == 'cdl':
                                        if wb0[s0].cell(row=r_cnt, column=c_cnt-1).value in ignore_cell and wb0[s0].cell(row=r_cnt, column=c_cnt+1).value:
                                                cdl_path = wb0[s0].cell(row=r_cnt, column=c_cnt+1).value
                                                if os.path.isdir(cdl_path):
                                                        for file in find_pattern('*.cdl', cdl_path):
                                                                #cdl.write(file+"\n")
                                                                cdl_list.append(file+"\n")
                                                        for file in find_pattern('*.spi', cdl_path):
                                                                #cdl.write(file+"\n")
                                                                cdl_list.append(file+"\n")
                                                elif glob.glob(cdl_path+'*.cdl') or glob.glob(cdl_path+'*.spi') or fnmatch.fnmatch(cdl_path, '*.spi') or fnmatch.fnmatch(cdl_path, '*.cdl'):
                                                        #cdl.write(cdl_path+"\n")
                                                        cdl_list.append(cdl_path+"\n")
                                #                cdl.write(wb0[s0].cell(row=r_cnt, column=c_cnt+1).value + "\n")
                                elif wb0[s0].cell(row=r_cnt, column=c_cnt).value == 'netwalker(pg)':
                                        if wb0[s0].cell(row=r_cnt, column=c_cnt-1).value in ignore_cell and wb0[s0].cell(row=r_cnt, column=c_cnt+1).value:           
                                                nwpg_path = wb0[s0].cell(row=r_cnt, column=c_cnt+1).value
                                                if os.path.isdir(nwpg_path):
                                                        for file in find_pattern('*.nw', nwpg_path):
                                                                #nwpg.write(file+"\n")
                                                                nwpg_list.append(file+"\n")
                                                elif glob.glob(nwpg_path+'*.nw') or fnmatch.fnmatch(nwpg_path, '*.nw'):
                                                        #nwpg.write(nwpg_path+"\n")
                                                        nwpg_list.append(nwpg_path+"\n")
                                #                nwpg.write(wb0[s0].cell(row=r_cnt, column=c_cnt+1).value + "\n")
                                                #wb1[sheet].cell(row=r_cnt, column=c_cnt).style = highlight
for x in sorted(set(gds_list)):
        gds.write(x)
for x in sorted(set(cdl_list)):
        cdl.write(x)
for x in sorted(set(nwpg_list)):
        nwpg.write(x)
#Analysing Verification sheet:

for s1 in wb1_sheets:
        if s1 in [GDSMERGE_1, CDLCAT]:
                continue
        else:

t1 =  time.time()
gds.close()
cdl.close()
nwpg.close()


print "Execution time: ", t1-t0

#eof
