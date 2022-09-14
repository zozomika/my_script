#!/common/appl/python/python-2.7.13-64bit/bin/python
# Function: Diff 2 workbook 
# Usage: script.py <old_excel> <new_excel>
# Highlight new update point of new excel file
import sys, time
from openpyxl import load_workbook, Workbook
from openpyxl.styles import PatternFill, Border, Side, Alignment, Protection, Font, colors, NamedStyle
from openpyxl.comments import Comment
from openpyxl.comments.comment_sheet import Properties
from subprocess import Popen, PIPE
import re

t0 = time.time()

wb0 = load_workbook(sys.argv[1])
wb1 = load_workbook(sys.argv[2])
output = Workbook()
# Check new added sheets or removed sheets
# Get only same sheets
wb0_sheets = wb0.sheetnames
wb1_sheets = wb1.sheetnames
for s1 in wb1_sheets:
    if s1 in wb0_sheets:
        pass
    else:
        print 'WARNING: This sheet is new added - %s' %s1
        wb1_sheets.remove(s1)
for s0 in wb0_sheets:
    if s0 in wb1_sheets:
        pass
    else:
        print 'ERROR: This sheet is removed - %s' %s0
        wb0_sheets.remove(s0)
##
highlight = NamedStyle('Highlight')
highlight.fill = PatternFill(fill_type='solid', start_color=colors.RED)
for sheet in wb1_sheets:
    max_row = max(wb1[sheet].max_row, wb0[sheet].max_row)
    max_col = max(wb1[sheet].max_column, wb0[sheet].max_column)
    for r_cnt in range(1,max_row+1):
        for c_cnt in range(1, max_col+1):
            print c_cnt, r_cnt
            if wb1[sheet].cell(row=r_cnt, column=c_cnt).value == wb0[sheet].cell(row=r_cnt, column=c_cnt).value:
                pass
            else:
                author = Popen('whoami',stdout=PIPE, shell=True).communicate()[0] #xem user
                wb1[sheet].cell(row=r_cnt, column=c_cnt).comment = Comment(wb0[sheet].cell(row=r_cnt, column=c_cnt).value, author)
                wb1[sheet].cell(row=r_cnt, column=c_cnt).style = highlight
wb1.save('diff.xlsx')

t1 =  time.time()
print "Execution time: ", t1-t0
