#!/common/appl/python/python-2.7.13-64bit/bin/python 
# -*- coding: utf-8 -*-
#.---------------- Author: HauHuynh
#.---------------- Usage : $> script.py <excel file> (*.xlsx)

import sys, time, datetime
from openpyxl import load_workbook, Workbook
#from openpyxl.styles import PatternFill, Border, Side, Alignment, Protection, Font, colors, NamedStyle
import re, os
reload(sys)
sys.setdefaultencoding('utf-8')

## Declaration
t0 = time.time()
#wb0 = load_workbook(sys.argv[1], data_only=True) # data_only => to get cell value not cell's content
wbo = load_workbook(r"https://renesasgroup.sharepoint.com/:x:/r/sites/RVC/Shared Documents/1. General Documents/010_ENG/190_Backend/RVBE document/02.Common/15.Elearning_Survey/collect_Survey_ELearning_BED_2022.xlsx", data_only=True)
#output = Workbook()
#stamp_out = '{:%y%m%d_%H%M}'.format(datetime.datetime.now())
output = open("output.txt", 'w')
wb0_sheets = wb0.sheetnames
#wb0_sheet = wb0['Output_sheet_results']

#def write_out (list_write_out, output):
#  for x in sorted(set(list_write_out)):
#    output.write(x)

# main script
print "[Info] Reading data from file %s   z.Z.Z..." % os.path.basename(sys.argv[1])

for s0 in wb0_sheets:
  output = open(s0+".txt", 'w')
  max_row = wb0[s0].max_row
  max_col = wb0[s0].max_column
  #print max_row
  #print max_col
  for r_cnt in range(1,max_row+1):
    for c_cnt in range(1, max_col+1):
        if wb0[s0].cell(row=r_cnt, column=c_cnt).value:
          output.write("%s " %wb0[s0].cell(row=r_cnt, column=c_cnt).value)
          #print wb0_sheet.cell(row=r_cnt, column=c_cnt).value
    output.write("\n")
  output.close()

t1 = time.time()
print "[Info] Finished in %f seconds " %(t1-t0)
wb0.close()

#eof
