#!/common/appl/python/python-2.7.13-64bit/bin/python                                                                      
import sys, time, datetime, csv
from openpyxl import load_workbook, Workbook
#from openpyxl.styles import PatternFill, Border, Side, Alignment, Protection, Font, colors, NamedStyle
#from openpyxl.comments import Comment
#from openpyxl.comments.comment_sheet import Properties
import re, os,  fnmatch
 
t0 = time.time()
t2 = time.time()
wb0 = load_workbook(sys.argv[1]) # EXCEL FILE
t3 = time.time()
print "Time to load workbook: %s" %(t3-t2)
output = Workbook()
# Setting
ignore_sheet = [] #["Rule","E3_History", "Delete(E3)","Old_V3M_History","hard_macro (MIPI)","D3_History", "Delete(D3)"]
ignore_cell = ["Use","use"]
 
# Initial declaration
 
stamp_out = '{:%y%m%d_%H%M}'.format(datetime.datetime.now()) #Ex: 171103_1234
# For excel
wb0_sheets = wb0.sheetnames
io_list  = []
pad_list = []
# --Output File
output_name =  "io_pad_info_" + stamp_out + ".txt"
output_file = open(output_name, 'w')                      
headline    = "%-3s %-55s %-30s %-30s \n" %("##", "INSTANT NAME","REF_NAME","PAD - REF_NAME")
#a, b, c = 1, 2.0, 'OK'
#print """a = %5s - %s
#b = %5s - %s
#c = %5s - %s\n""" %(a, type(a), b, type(b), c, type(c))

output_file.write(headline)
i = 1
for sheet in wb0_sheets:
    max_row = wb0[sheet].max_row
    max_col = wb0[sheet].max_column    
    for r_cnt in range(1,max_row+1):
        #for c_cnt in range(1, max_col+1):
            #print c_cnt, r_cnt
            #if wb0[sheet].cell(row=r_cnt, column=c_cnt).value == "BL_CORNER":
                #print r_cnt, c_cnt
            io_name = wb0[sheet].cell(row=r_cnt, column=32).value
            io_ref_name = wb0[sheet].cell(row=r_cnt, column=33).value
            pad_ref_name = wb0[sheet].cell(row=r_cnt, column=22).value
            if r_cnt >= 44 and io_name: 
                output_file.write("%-3s %-55s %-30s %-30s \n" %(i, io_name, io_ref_name, pad_ref_name))
                i += 1
            else:
               # author = Popen('whoami',stdout=PIPE, shell=True).communicate()[0] #xem user
               # wb1[sheet].cell(row=r_cnt, column=c_cnt).comment = Comment(wb0[sheet].cell(row=r_cnt, column=c_cnt).value,
               # wb1[sheet].cell(row=r_cnt, column=c_cnt).style = highlight
                pass
output_file.close()
##EOF
