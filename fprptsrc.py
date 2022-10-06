"""Utility functions for copying and archiving files and directory trees.                                                                                                                                                                    
 
  Project: RV28F
  File name: wstatsrc.py
  Author: RVC/BE31/Hau Huynh
 ------------------------------------------------------------------------
  
  History
  v1r0 : Nov.11 2019  by hauhuynh/RVC/BE31
       - Get text file and importing to excel file
  v1r1 : Nov.18 2019  by hauhuynh/RVC/BE31
       - Get wire statistic and fill-in excel table
       - Coloring and mapping table content
  v1r2 : Nov.21 2019  by hauhuynh/RVC/BE31
       - Update xcl_wstat: get IP name, IP Pin attribute (name, layer, width)
       - Update excel format: store new updated information
"""

import re, os, fnmatch
 
#import function from lib source
from subprocess import Popen, PIPE
from openpyxl import load_workbook, Workbook
from openpyxl.styles import PatternFill, Border, Side, Alignment, Protection, Font, colors
from openpyxl.drawing.image import Image
from openpyxl.utils import get_column_letter, column_index_from_string
from datetime import datetime
#from copy import copy, deepcopy
import shutil 

def check_ln_pattern(line, pattern):
  if not re.match (r'^\#.*$', line):
    if re.match (pattern, line):
      return 1
    else :
      return 0

def check_ln_pattern2(line, pattern):
  if re.search (pattern, line):
    return 1
  else :
    return 0

def check_ln_pattern_relax(line, pattern):
  if re.match (pattern, line):
    return 1
  else :
    return 0

def read_txt (_input):
  rpt = open(_input, 'r')
  return rpt

def find_all(name, path):
#get all matches exactly
    result = []
    for root, dirs, files in os.walk(path):
        if name in files:
            result.append(os.path.join(root, name))
    return result

def find_pattern (pattern, path):
#get match by patterns:
    result = []
    for root, dirs, files in os.walk(path):
        for name in files:
            #if str.lower(name) in [x.lower() for x in files]:
            if fnmatch.fnmatch(name, pattern) :
                result.append(os.path.join(root, name))
    return result


def clone_wb2 (fr_wb, to_wb):
  print ("[INFO] Clone excel sheet: %s" %to_wb)
  shutil.copy(fr_wb, to_wb)

def find_file(pattern, path):
  # get 1st match
  for root, dirs, files in os.walk(path):
    for f in files:
      if fnmatch.fnmatch(f, pattern):
        return os.path.join(root, f)
  

def make_wb ( _wb, _ws ):
  wb            = Workbook()
  ws1           = wb.active
  if _ws != "":
    ws1.title     = _ws
  if _wb != "":
    wb.save(_wb)
    return _wb

def find_evidence(str_arr, output_str , current_pos, end_loop_str):
  enum = 0
  while 1:
    output_str = output_str + str_arr[current_pos+enum]
    enum += 1
    if check_ln_pattern(str_arr[current_pos+enum], end_loop_str):
      output_str = output_str + str_arr[current_pos+enum]
      break
  return output_str

def find_evidence_relax(str_arr, output_str , current_pos, end_loop_str):
  enum = 0
  while 1:
    output_str = output_str + str_arr[current_pos+enum]
    enum += 1
    if check_ln_pattern_relax(str_arr[current_pos+enum], end_loop_str):
      output_str = output_str + str_arr[current_pos+enum]
      break
  return output_str

def find_image (_dir, img_name):
  result = ""
  try:
    result = find_file(img_name, _dir)
    print ("[INFO] Found %s" %result)
  except:
    print ("[Error] Not found %s image" %img_name)
  return result

def find_images (_dir, img_name):
  result = ""
  try:
    result = find_pattern(img_name, _dir)
    if result == []:
      print ("[Error] Not found %s image" %img_name)
    else:
      print ("[INFO] Found %s" %result)
  except:
    pass
  return result

def what_time_is_it ():
  now = datetime.now()
  time = now.strftime("%m/%d/%Y, %H:%M:%S")
  return time

def what_date_is_it ():
  now = datetime.now()
  date = now.strftime("%m/%d/%Y")
  return date

def user_name ():
  usr = Popen(['whoami'], universal_newlines=True, stdout=PIPE).communicate()[0]
  return usr.replace('\n','')

def sheet_content_Cover     (_wb, _ws, _user):
  all_ws = _wb.worksheets
  #Define cell's attribute
  cell = {'C3' : _user}
  for ws in all_ws:
    if ws.sheet_state == "visible":
      if ws.title == _ws:
        print ("[INFO] Start processing \"%s\" sheet" %ws.title)
        for k,v in cell.items():
          print ("Cell %s" %k)
          if v == "":
            ws[k].value = "[Error] not found evidence"
          else:
            ws[k].value = "by " + v + " at " + what_time_is_it()
      else:
        pass

def sheet_content_Floorplan (_wb, _ws, _lib, _lib_evidence, _chipsize, _chipsize_evidence,
                                       _netlist, _netlist_evidence, _iopad, _iopad_evidence, 
                                       _upf, _upf_evidence, _tcd_evidence, _tcd,
                                      _track, _row,
                                      _tapcap,
                                      _ser,
                                      _fv, _tie_evidence,
                                      _rep, _gr_evidence,
                                      _RB_evidence,
                                      _floorplan_evidence):
  all_ws = _wb.worksheets
  v_arr  = []
  #all_ws = _wb.active
  #Define cell's attribute
  cell = {'G12' : _lib, 'M12' : _lib_evidence, 'G13' : _chipsize, 'G14' : _netlist, 'M14' : _netlist_evidence,'G19' : _iopad, 'M19' : _iopad_evidence, 
          'G20' : _upf, 'M20' : _upf_evidence, 'G21' : _iopad, 'M21' : _iopad_evidence, 'M22': _tcd_evidence, 'G24' : _track, 'G27' : _row,
          'G29' : _tapcap, 'M32' : _ser, 'M33' : _ser, 'M34' : _ser, 'M67' : _fv, 'M13' : _chipsize_evidence, 'M18' : _tie_evidence,
          'N22' : _tcd, 'M28' : _rep, 'M31' : _gr_evidence, 'M71' : _RB_evidence, 'M70' : _floorplan_evidence}
  for ws in all_ws:
    if ws.sheet_state == "visible":
      if ws.title == _ws:
        print ("[INFO] Start processing \"%s\" sheet" %ws.title)
        for k,v in cell.items():
          print ("Cell %s" %k)
          r = ws[k].row
          c = ws[k].column
          c_index = column_index_from_string(c)
          j = "J" + str(r) ; # Judgment cell is in column J
          jv = user_name().split('_')[0].upper() + "\n" + what_date_is_it()
          ws[j].value = jv
          if v == "":
            ws[k].value = "[Error] not found evidence"
          else:
            try:
              ws[k].value = v
              img = Image(v)
              img.height = 968
              img.width  = 1920
              ws.row_dimensions[ws[k].row].height = 800
              ws.add_image(img,k)
            except:
              try:
                if isinstance(v,list):
                  for va in v:
                    #print(str(chr(c_tmp)))
                    k = c + str(r)
                    img = Image(va)
                    img.height = 968
                    img.width  = 1920
                    ws.row_dimensions[ws[k].row].height = 800
                    ws.column_dimensions[ws[k].column].width  = 340
                    ws.add_image(img,k)
                    c_index = int(c_index) + 1
                    c = get_column_letter(c_index)
              except:
                pass
        #ws[cell['Lib']].value = _lib
        #ws[cell['Lib_evidence']].value = _lib_evidence
        #ws[cell['Chipsize']].value = _chipsize
  

def xcl_fpstat ( fprpt, output ):
  "A function to generate excel report using input data is wire's info report"
  print ("[INFO] Start making Excel report...")
  num           = 0
  num1          = 0
  num2          = 0
  info_lib      = ""
  info_chipsize = ""
  info_netlist  = ""
  info_iopad    = ""
  info_upf      = ""
  info_tcd      = ""
  info_row      = ""
  info_track    = ""
  info_ser      = ""
  info_fv       = ""
  end_flag      = 0
  my_rpt        = read_txt (fprpt)
  myrpt_path    = os.path.abspath(fprpt)
  myrpt_dir     = os.path.dirname(myrpt_path)
  myrpt_wdir    = os.path.dirname(myrpt_dir)
  ln_rpt = my_rpt.readlines()
  clone_wb2 ("/svhome/VCF_RVC_OffSite/hau_p2f2_vf/U2A6-BEDQuality-0001-01-Floorplan_CN_f5090_0013d.xlsx", output)
  #print ("[INFO] Reading data from file %s" % myrpt_path)
  my_xcl = load_workbook(filename=output)
  lib_rpt      = find_file('*report_ref_libs', os.path.dirname(os.path.dirname(myrpt_dir)+'/reports'))
  init_file = os.path.dirname(myrpt_dir) + '/icc2_scripts/100_init_design.tcl'
   
  try:
    my_lib_file     = read_txt(lib_rpt)
    print ("[INFO] Processing lib file %s"%lib_rpt)
    my_lib_evidence = os.path.dirname(myrpt_dir) + '/reports/report_lib_cell_purpose.rpt' + '\n'
    for lline in my_lib_file.readlines():
      my_lib_evidence = my_lib_evidence + lline
  except:
    print ("[Error] Not found %s" %lib_rpt)
  #my_lib_evidence[0] = os.path.dirname(myrpt_dir) + '/reports/report_lib_cell_purpose.rpt' + '\n' + my_lib_evidence[0]

  print ("[INFO] Processing log file %s" %fprpt)
  for line in ln_rpt: 
    # Processing on fprpt
    line = line.strip()
    line = line + '\n'
    if not end_flag:
      num += 1
      line = line.replace('\r\n','\n')
      if check_ln_pattern(line, r'.*source.*100_init_design_.*'):
        print(line)
        line_arr = line.split()
        init_file = os.path.dirname(myrpt_dir) + '/' + line.split()[len( line_arr)-1]
      elif check_ln_pattern(line, r'^RM-info : Running script.*icc2_setup.*'):
        setup_file = line.split()[4]
      elif check_ln_pattern(line, r'^RM-info : Running script.*common_setup.*'):
        common_setup_file = line.split()[4]
        #setup_file_path = os.path.dirname(myrpt_path) + '/' + os.path.basename(setup_file)
      elif check_ln_pattern(line, r'.*ICC2_LIB_PATH .*'):
        info_lib = info_lib + line
        #G12: LIB
        info_lib = find_evidence_relax(ln_rpt, info_lib, num, '.*REFERENCE_LIBRARY.*')
      elif check_ln_pattern(line, r'set VERILOG_NETLIST_FILE.*'):
        #G14: NETLIST
        info_netlist = line
      elif check_ln_pattern(line, r'START_CMD: read_verilog.*'):
        netlist_evidence = line
        #M14: NETLIST evidence
        netlist_evidence = find_evidence(ln_rpt, netlist_evidence, num, r'.*END_CMD: read_verilog.*')
      elif check_ln_pattern2(line, 'IO Fill & ESD'):
        iopad_evidence = line
        #M19: IOPAD evidence                        
        iopad_evidence = find_evidence(ln_rpt, iopad_evidence, num, r'END_CMD: read_def.*')
      elif check_ln_pattern(line, '.*Loading DEF file.*PAD.*'):
        iopad_evidence = iopad_evidence + '\n' + line
        iopad_evidence = find_evidence(ln_rpt, iopad_evidence, num, r'END_CMD: read_def.*')
      elif check_ln_pattern(line,r'set UPF.*'):
        #G20: UPF
        info_upf = info_upf + line
        upf_evidence = ""
      elif check_ln_pattern(line,r'source \$UPF.*'):
        #M20: UPF evidence
        upf_evidence = upf_evidence + line
      elif check_ln_pattern(line,r'commit_upf'):
        upf_evidence = upf_evidence + line
        upf_evidence = find_evidence(ln_rpt, upf_evidence, num, r'[01]')
      elif check_ln_pattern2(line,'TCD Dummy'):
        #M22: TCD evidence
        tcd_evidence = line
        tcd_evidence = find_evidence(ln_rpt, tcd_evidence, num, r'END_CMD: read_def.*')
      elif check_ln_pattern(line, r'.*u2a6_lib:100_init_design.design.*'):
        end_flag = 1

  my_init_file = read_txt(init_file)
  init_rpt = my_init_file.readlines()
  print ("[INFO] Processing init file %s" %init_file)
  for iline in init_rpt:
    iline = iline.strip()
    iline = iline + '\n'
    num2 += 1
    if check_ln_pattern(iline, r'set CORE_SIDE_LENGTH.*'):
    #G13: CHIP SIZE
      info_chipsize = info_chipsize + iline
    elif check_ln_pattern(iline, r'.*initialize_floorplan.*'):
      info_chipsize = info_chipsize + iline
      info_chipsize = find_evidence(init_rpt, info_chipsize, num2, r'.*coincident_boundary.*')
    elif check_ln_pattern(iline, r'create_track.*'):
      #G27
      #iline_tmp = iline.strip()
      #if not check_ln_pattern2(info_track, "%s" %iline_tmp):
      ###BUG: can not filter info_track
        if iline not in info_track:
          info_track = info_track + iline
    elif check_ln_pattern(iline, r'foreach.*row_boundary.*'):
      info_row = info_row + iline
      info_row = find_evidence(init_rpt, info_row, num2, r'}')
    elif check_ln_pattern(iline, r'source.*endcap_tap.*'):
      info_tapcap = "SCRIPT\n" + iline
    elif check_ln_pattern(iline, r'set DEF_IO.*'):
      #G19
      info_iopad = info_iopad + iline
    elif check_ln_pattern(iline, r'set DEF_PAD.*'):
      info_iopad = info_iopad + iline
  
  #G41-G42-G43: IP Orientation & location
  ser_file = os.path.dirname(myrpt_dir) + '/CMD_icc2/ser_chk/total_report.rpt'
  try:
    my_ser_file = read_txt(ser_file)
    print ("[INFO] Processing Sercheck file %s" %ser_file)
    ser_rpt = my_ser_file.readlines()
    for sline in ser_rpt:
      sline = sline.strip()
      sline = sline + '\n'
      info_ser = info_ser + sline
  except:
    print ("[Error] Not found %s" %ser_file)

  #M67: FV
  fv_file = os.path.dirname(myrpt_dir) + '/FV/conformal.log'
  try:
    my_fv_file = read_txt(fv_file)
    print ("[INFO] Processing Conformal file %s" %fv_file)
    #info_fv = ":\n%s\n" %fv_file
    fv_rpt = my_fv_file.readlines()
    for fline in fv_rpt:
      num1 += 1
      fline = fline.strip()
      fline = fline + '\n'
      if check_ln_pattern2(fline, 'report compare data -class nonequivalent'):
        info_fv = info_fv + fline
        info_fv = find_evidence(fv_rpt,info_fv,num1,r'.*report statistics')
  except:
    print ("[Error] Not found %s" %fv_file)
  
  #M13: Chipsize evidence
  chip_evidence = find_image(os.path.dirname(myrpt_dir)+'/results','*chipsize*')
  #M18: Tiecell evidence
  tie_evidence = find_image(os.path.dirname(myrpt_dir)+'/results','*tiecell*')
  #M24: TCD evidence
  tcd_img = find_image(os.path.dirname(myrpt_dir)+'/results','*tcdcell*')
  #M34: 5V Repeater evidence
  rep_img = find_images(os.path.dirname(myrpt_dir)+'/results','*5V_REP*')
  #M39: Guardring
  GR_img = find_images(os.path.dirname(myrpt_dir)+'/results','*guardring*')
  #M39: Guardring
  RB_img = find_images(os.path.dirname(myrpt_dir)+'/results','*analogRB*')
  #M59: Guardring
  FP_img = find_images(os.path.dirname(myrpt_dir)+'/results','*floorplan*')
  #if os.path.exist(setup_file):
  #   my_setup = read_txt(setup_file)
  #   setup_lines = my_setup.readlines()
  sheet_content_Floorplan (my_xcl, 'Floorplan', info_lib, my_lib_evidence,
                                                info_chipsize, chip_evidence,
                                                info_netlist, netlist_evidence,
                                                info_iopad, iopad_evidence,
                                                info_upf, upf_evidence,
                                                tcd_evidence, tcd_img,
                                                info_track,
                                                info_row,
                                                info_tapcap,
                                                info_ser,
                                                info_fv,
                                                tie_evidence, rep_img,
                                                GR_img, RB_img,
                                                FP_img)
  sheet_content_Cover (my_xcl, 'Cover', user_name())
  my_xcl.save(output)
  print ("Finished!")
#eof
